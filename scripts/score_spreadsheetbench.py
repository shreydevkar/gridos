"""Score GridOS outputs against SpreadsheetBench Verified-400 goldens.

Reads the produced xlsx files from
    {dataset}/outputs/{setting}_{model}/1_{id}_output.xlsx
compares cells in answer_position against
    {dataset}/spreadsheet/{id}/1_{id}_golden.xlsx
and emits two metrics in the spirit of SpreadsheetBench's eval:

  hard_restriction (per question)  = 1.0 if EVERY cell in answer_position
                                     matches; 0.0 otherwise.
  soft_restriction (per question)  = matched_cells / total_cells.

Reported aggregates:
  - count of questions actually scored (skipping ones that didn't produce output)
  - mean hard score (= % of questions where the answer is fully correct)
  - mean soft score (= average cell-level correctness)
  - per-instruction-type breakdown so Cell-Level vs Sheet-Level wins/losses
    are visible

Usage:
    python scripts/score_spreadsheetbench.py
    python scripts/score_spreadsheetbench.py --limit 25 --setting single --model gridos
"""
import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Optional

import openpyxl


DEFAULT_DATASET = (
    Path(__file__).resolve().parent.parent
    / "bench"
    / "spreadsheetbench"
    / "spreadsheetbench_verified_400"
)


def parse_args():
    p = argparse.ArgumentParser(description="Score SpreadsheetBench outputs.")
    p.add_argument("--dataset", default=str(DEFAULT_DATASET))
    p.add_argument("--setting", default="single")
    p.add_argument("--model", default="gridos")
    p.add_argument("--start", type=int, default=0,
                   help="Skip the first N questions in dataset.json before scoring.")
    p.add_argument("--limit", type=int, default=0,
                   help="Score only the first N questions (0 = all).")
    p.add_argument("--out", default=None,
                   help="Where to write the per-question JSON. "
                        "Default: {dataset}/outputs/score_{setting}_{model}.json")
    return p.parse_args()


def _col_letter(col_num: int) -> str:
    s = ""
    while col_num > 0:
        col_num, rem = divmod(col_num - 1, 26)
        s = chr(65 + rem) + s
    return s


def _col_num(col_letter: str) -> int:
    n = 0
    for c in col_letter.upper():
        n = n * 26 + (ord(c) - 64)
    return n


def _gen_cells(answer_position: str) -> list[tuple[Optional[str], str]]:
    """Expand answer_position (Excel-style) into a flat list of (sheet, A1) tuples.
    Handles multi-range comma syntax, sheet-qualified prefixes, and ranges."""
    out = []
    for sub in answer_position.split(","):
        if "!" in sub:
            sheet_part, cell_part = sub.split("!", 1)
            sheet_name = sheet_part.strip().lstrip("'").rstrip("'")
        else:
            sheet_name, cell_part = None, sub
        cell_part = cell_part.strip().lstrip("'").rstrip("'")
        if ":" in cell_part:
            a, b = cell_part.split(":")
            ma = re.match(r"([A-Z]+)(\d+)", a, re.I)
            mb = re.match(r"([A-Z]+)(\d+)", b, re.I)
            if not (ma and mb):
                continue
            c1, c2 = _col_num(ma.group(1)), _col_num(mb.group(1))
            r1, r2 = int(ma.group(2)), int(mb.group(2))
            for r in range(min(r1, r2), max(r1, r2) + 1):
                for c in range(min(c1, c2), max(c1, c2) + 1):
                    out.append((sheet_name, f"{_col_letter(c)}{r}"))
        else:
            out.append((sheet_name, cell_part))
    return out


def _values_match(g, o) -> bool:
    """SpreadsheetBench-style cell comparison.
    - Numeric: round to 2 decimals (matches their evaluation.py).
    - Datetime: compare on calendar (date) only.
    - Strings (numeric-looking): try float coercion both sides.
    - Empty cells: None / '' are equivalent.
    """
    if (g is None or g == "") and (o is None or o == ""):
        return True
    # Datetime — compare date portion if both look like dates
    import datetime
    if isinstance(g, datetime.datetime) or isinstance(o, datetime.datetime):
        if isinstance(g, datetime.datetime) and isinstance(o, datetime.datetime):
            return g.date() == o.date() and g.time() == o.time()
        # one is a datetime, the other isn't — try to coerce the string side
        return False
    # Try float comparison if either is numeric
    try:
        gf = float(g) if not isinstance(g, bool) else (1 if g else 0)
        of = float(o) if not isinstance(o, bool) else (1 if o else 0)
        return round(gf, 2) == round(of, 2)
    except (ValueError, TypeError):
        pass
    # Fallback: case-insensitive string equality
    if isinstance(g, str) or isinstance(o, str):
        gs = "" if g is None else str(g)
        os_ = "" if o is None else str(o)
        return gs.strip().lower() == os_.strip().lower()
    return g == o


def score_one(out_xlsx: Path, gold_xlsx: Path, answer_position: str) -> dict:
    if not out_xlsx.exists():
        return {"scored": False, "reason": "no output xlsx"}
    try:
        wb_o = openpyxl.load_workbook(out_xlsx, data_only=True)
        wb_g = openpyxl.load_workbook(gold_xlsx, data_only=True)
    except Exception as e:
        return {"scored": False, "reason": f"load error: {type(e).__name__}: {e}"}

    cells = _gen_cells(answer_position)
    if not cells:
        return {"scored": False, "reason": f"empty answer_position: {answer_position!r}"}

    matches = 0
    total = 0
    for sheet_name, a1 in cells:
        # Resolve target sheet on each side (case-insensitive)
        if sheet_name:
            sg = next((wb_g[n] for n in wb_g.sheetnames if n.lower() == sheet_name.lower()), None)
            so = next((wb_o[n] for n in wb_o.sheetnames if n.lower() == sheet_name.lower()), None)
        else:
            sg = wb_g[wb_g.sheetnames[0]]
            so = wb_o[wb_o.sheetnames[0]] if wb_o.sheetnames else None
        total += 1
        if sg is None or so is None:
            continue
        gv = sg[a1].value
        ov = so[a1].value
        if _values_match(gv, ov):
            matches += 1
    return {
        "scored": True,
        "total": total,
        "matched": matches,
        "soft": matches / total if total else 0.0,
        "hard": 1.0 if matches == total and total > 0 else 0.0,
    }


def main() -> int:
    args = parse_args()
    dataset_dir = Path(args.dataset).resolve()
    if not (dataset_dir / "dataset.json").exists():
        print(f"[error] dataset.json not found at {dataset_dir}", file=sys.stderr)
        return 2

    with open(dataset_dir / "dataset.json", "r", encoding="utf-8") as fp:
        dataset = json.load(fp)
    if args.start:
        dataset = dataset[args.start:]
    if args.limit > 0:
        dataset = dataset[: args.limit]

    out_dir = dataset_dir / "outputs" / f"{args.setting}_{args.model}"
    if not out_dir.is_dir():
        print(f"[error] no outputs at {out_dir}", file=sys.stderr)
        return 2

    per_q = []
    by_type = defaultdict(lambda: {"n": 0, "hard": 0.0, "soft": 0.0})
    n_scored = 0
    sum_hard = sum_soft = 0.0

    for item in dataset:
        qid = str(item["id"])
        out_xlsx = out_dir / f"1_{qid}_output.xlsx"
        gold_xlsx = dataset_dir / "spreadsheet" / qid / f"1_{qid}_golden.xlsx"
        result = score_one(out_xlsx, gold_xlsx, item["answer_position"])
        result["id"] = qid
        result["instruction_type"] = item.get("instruction_type", "?")
        per_q.append(result)
        if result["scored"]:
            n_scored += 1
            sum_hard += result["hard"]
            sum_soft += result["soft"]
            t = item.get("instruction_type", "?")
            by_type[t]["n"] += 1
            by_type[t]["hard"] += result["hard"]
            by_type[t]["soft"] += result["soft"]

    out_json = args.out or (dataset_dir / "outputs" / f"score_{args.setting}_{args.model}.json")
    Path(out_json).parent.mkdir(parents=True, exist_ok=True)
    with open(out_json, "w", encoding="utf-8") as fp:
        json.dump({
            "n_total": len(dataset),
            "n_scored": n_scored,
            "n_skipped": len(dataset) - n_scored,
            "mean_hard": (sum_hard / n_scored) if n_scored else 0.0,
            "mean_soft": (sum_soft / n_scored) if n_scored else 0.0,
            "per_instruction_type": {
                t: {
                    "n": v["n"],
                    "mean_hard": v["hard"] / v["n"] if v["n"] else 0.0,
                    "mean_soft": v["soft"] / v["n"] if v["n"] else 0.0,
                }
                for t, v in by_type.items()
            },
            "per_question": per_q,
        }, fp, indent=2, default=str)

    # Console summary
    print(f"=== SpreadsheetBench V2 — {args.setting} / {args.model} ===")
    print(f"Questions in run:  {len(dataset)}")
    print(f"Scored:            {n_scored}")
    print(f"Skipped (no out):  {len(dataset) - n_scored}")
    if n_scored:
        print(f"Mean HARD score:   {sum_hard / n_scored * 100:.1f}%   (% of questions with EVERY cell correct)")
        print(f"Mean SOFT score:   {sum_soft / n_scored * 100:.1f}%   (mean cell-level correctness)")
    print()
    print("By instruction type:")
    for t, v in sorted(by_type.items()):
        if v["n"]:
            print(f"  {t:<26}  n={v['n']:>2}  hard={v['hard']/v['n']*100:5.1f}%  soft={v['soft']/v['n']*100:5.1f}%")
    print()
    print(f"Per-question detail: {out_json}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
