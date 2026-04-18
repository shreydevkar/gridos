import json
import os
import random
import re
import time
from collections import OrderedDict
from contextvars import ContextVar
from datetime import datetime, timezone
from pathlib import Path
from threading import Lock
from typing import Any, Dict, List, Optional

from dotenv import load_dotenv
from fastapi import Body, FastAPI, File, Header, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from agents import load_agents
from core.engine import GridOSKernel
from core.functions import FormulaEvaluator
from core.macros import MacroError, compile_macro
from core.models import AgentIntent, WriteResponse
from core.providers import (
    AnthropicProvider,
    GeminiProvider,
    GroqProvider,
    MODEL_CATALOG,
    OpenRouterProvider,
    Provider,
    ProviderAuthError,
    ProviderTransientError,
    default_model_id,
    get_model_entry,
)
from core.utils import a1_to_coords


load_dotenv()
TELEMETRY_PATH = Path("telemetry_log.json")
MAX_CHAIN_ITERATIONS = 10

DATA_DIR = Path("data")
TEMPLATES_DIR = DATA_DIR / "templates"
MACROS_PATH = DATA_DIR / "user_macros.json"
HERO_TOOLS_PATH = DATA_DIR / "hero_tools.json"
API_KEYS_PATH = DATA_DIR / "api_keys.json"

HERO_TOOLS_CATALOG = [
    {
        "id": "web_search",
        "display_name": "Web Search",
        "description": "(placeholder) Advises the agent that live web lookups are available. Actual fetching is not wired up.",
    },
    {
        "id": "live_data",
        "display_name": "Live Data Puller",
        "description": "(placeholder) Advises the agent that external API/data feeds are available. Actual fetching is not wired up.",
    },
]

app = FastAPI(title="GridOS - Agentic Workbook")
AGENTS = load_agents()

# Per-request kernel resolution. In OSS there's exactly one workbook state and
# _default_kernel is it. In SaaS we keep a pool of GridOSKernel instances keyed
# by (user_id, workbook_id); each request binds to the right one via the
# current_kernel_dep dependency (below), which sets _current_kernel ContextVar
# before the endpoint body runs. The `kernel` name here is a thin proxy — every
# `kernel.X` access in existing code automatically routes to the request-scoped
# kernel, so we don't have to thread `k` through hundreds of call sites.
_default_kernel = GridOSKernel()
_current_kernel: ContextVar[Optional[GridOSKernel]] = ContextVar("gridos_current_kernel", default=None)
_kernel_pool: "OrderedDict[tuple[str, str], GridOSKernel]" = OrderedDict()
_kernel_pool_lock = Lock()
# Cap pool size so an abusive bot can't OOM the process. LRU-evict the oldest
# entries past the cap. Render free = 512 MB → 64 kernels gives us comfortable
# headroom. Evicted users' next request just lazy-loads from Supabase again.
KERNEL_POOL_MAX = 64


class _KernelProxy:
    """Reads the per-request kernel from the ContextVar.

    OSS: ContextVar may be unset (e.g. code path that doesn't go through the
    dep); fall through to _default_kernel — behavior matches pre-refactor.

    SaaS: any access without a resolved ContextVar is a bug — it would silently
    read/write the wrong user's workbook state (the exact class of bug this
    refactor prevents). Raise loudly so we catch missed Depends() wiring in CI
    instead of corrupting a user's data in prod."""

    def __getattr__(self, name: str):
        k = _current_kernel.get()
        if k is None:
            if cloud_config.SAAS_MODE:
                raise RuntimeError(
                    f"kernel.{name} accessed outside a request-scoped kernel in SaaS mode. "
                    "The endpoint is missing `Depends(current_kernel_dep)` on its signature."
                )
            k = _default_kernel
        return getattr(k, name)


kernel = _KernelProxy()

USER_MACROS: list[dict] = []
HERO_TOOLS_STATE: dict[str, bool] = {t["id"]: False for t in HERO_TOOLS_CATALOG}

os.makedirs("static", exist_ok=True)
TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
app.mount("/static", StaticFiles(directory="static"), name="static")

# Cloud (SaaS) router is always mounted — it exposes /cloud/status which the
# frontend reads on bootstrap to decide whether to show login/billing UI. Real
# SaaS features are gated on config.SAAS_MODE and attached in later phases.
from cloud import config as cloud_config  # noqa: E402
from cloud import usage as cloud_usage  # noqa: E402
from cloud.auth import AuthUser, require_user  # noqa: E402
from cloud.status import router as cloud_status_router  # noqa: E402
from core.workbook_store import FileWorkbookStore, WorkbookScope, WorkbookStore  # noqa: E402
from fastapi import Depends  # noqa: E402

app.include_router(cloud_status_router)

# Persistence seam. In OSS mode we keep using the flat-file store so behavior
# is bit-identical. In SaaS mode we try Supabase and fall back to the file
# store with a loud warning if credentials are missing — a misconfigured SaaS
# deploy should surface as 503s on save/load, not crash at startup.
workbook_store: WorkbookStore
if cloud_config.SAAS_MODE and cloud_config.SAAS_FEATURES["cloud_storage"].enabled:
    from cloud.supabase_store import SupabaseWorkbookStore  # noqa: E402

    workbook_store = SupabaseWorkbookStore(
        url=cloud_config.SUPABASE_URL,
        key=cloud_config.SUPABASE_KEY,
    )
elif cloud_config.SAAS_MODE:
    print("[cloud] SAAS_MODE=true but SUPABASE_URL/KEY missing — /system/save and /system/load will return 503.")
    workbook_store = FileWorkbookStore()
else:
    workbook_store = FileWorkbookStore()


# ---------- Library persistence ----------


def _builtin_primitive_names() -> list[str]:
    # _default_kernel's evaluator is always populated and identical across
    # kernels for built-ins (macros are added by _register_macro which walks
    # every kernel). Using the default keeps this callable outside a request
    # scope (e.g. /tools/list which doesn't touch workbook state).
    return sorted(_default_kernel.evaluator.registry.keys())


def _macro_names() -> set[str]:
    return {m["name"].upper() for m in USER_MACROS}


def _register_macro_into(k: GridOSKernel, spec: dict) -> None:
    """Compile and register a macro into a single kernel's evaluator."""
    macro_name = spec["name"].upper()
    # Exclude any previously-registered version of this macro from the primitive pool
    # so macros can be updated safely and so a macro can't (accidentally) recurse into itself.
    primitive_registry = {
        name: fn for name, fn in k.evaluator.registry.items() if name.upper() != macro_name
    }
    fn = compile_macro(
        name=spec["name"],
        params=spec.get("params", []),
        body=spec["body"],
        registry=primitive_registry,
    )
    k.evaluator.register_custom(macro_name, fn)


def _register_macro(spec: dict) -> None:
    """Register a macro into _default_kernel AND every live per-user kernel in
    the pool. At startup the pool is empty so only _default_kernel is touched;
    at runtime this propagates newly-saved macros to every active session."""
    _register_macro_into(_default_kernel, spec)
    with _kernel_pool_lock:
        pool_snapshot = list(_kernel_pool.values())
    for k in pool_snapshot:
        try:
            _register_macro_into(k, spec)
        except MacroError:
            # Already validated against _default_kernel — per-kernel failures
            # would indicate divergent registry state, which we silently skip
            # so one bad kernel doesn't block the broadcast.
            continue


def _unregister_macro(upper_name: str) -> None:
    """Drop a macro from every live kernel so delete takes effect immediately
    across all sessions — not just the one that issued the request."""
    _default_kernel.evaluator.registry.pop(upper_name, None)
    with _kernel_pool_lock:
        for k in _kernel_pool.values():
            k.evaluator.registry.pop(upper_name, None)


def _load_user_macros() -> None:
    if not MACROS_PATH.exists():
        return
    try:
        raw = json.loads(MACROS_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return
    if not isinstance(raw, list):
        return
    for spec in raw:
        try:
            _register_macro(spec)
        except MacroError:
            # Drop invalid stored macros silently; they're user-authored and may
            # predate schema changes. They'll reappear on the next successful save.
            continue
        USER_MACROS.append({
            "name": spec["name"].upper(),
            "description": spec.get("description", ""),
            "params": [p.upper() for p in spec.get("params", [])],
            "body": spec["body"],
        })


def _persist_user_macros() -> None:
    MACROS_PATH.write_text(json.dumps(USER_MACROS, indent=2), encoding="utf-8")


def _load_hero_tools() -> None:
    if not HERO_TOOLS_PATH.exists():
        return
    try:
        raw = json.loads(HERO_TOOLS_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return
    if not isinstance(raw, dict):
        return
    for tool in HERO_TOOLS_CATALOG:
        HERO_TOOLS_STATE[tool["id"]] = bool(raw.get(tool["id"], False))


def _persist_hero_tools() -> None:
    HERO_TOOLS_PATH.write_text(json.dumps(HERO_TOOLS_STATE, indent=2), encoding="utf-8")


_load_user_macros()
_load_hero_tools()


# ---------- Per-request kernel resolution ----------
# Defined here (not later next to endpoint handlers) so every @app.X endpoint
# below can reference current_kernel_dep as a Depends(). All of these need to
# exist before Python evaluates the endpoint function signatures.


def _scope_for(user: AuthUser, workbook_id: Optional[str] = None) -> WorkbookScope:
    """Map auth context to a WorkbookScope.

    OSS: single-user, single-workbook — user_id=None, workbook_id="default"
        (legacy `system_state.gridos` file). An explicit workbook_id is
        ignored.
    SaaS: multi-workbook. Callers pass the active workbook's id; we fall
        back to user.id for back-compat with the original one-workbook-per
        -user behavior (used by save/load when the frontend doesn't pass
        the id yet).
    """
    if not cloud_config.SAAS_MODE:
        return WorkbookScope(user_id=None, workbook_id="default")
    return WorkbookScope(user_id=user.id, workbook_id=workbook_id or user.id)


def _register_macros_into_fresh(k: GridOSKernel) -> None:
    """Replay the global USER_MACROS list onto a freshly-created kernel.
    Called exactly once per kernel, when it enters the pool."""
    for spec in USER_MACROS:
        try:
            _register_macro_into(k, spec)
        except MacroError:
            continue


def _kernel_for_scope(scope: WorkbookScope) -> GridOSKernel:
    """Resolve (or create) the GridOSKernel for this scope.

    OSS: always returns _default_kernel — preserves pre-refactor behavior.

    SaaS: returns (creating + lazy-loading from Supabase if needed) the kernel
    for this (user, workbook). LRU-evicts the oldest entry when past the cap.
    Thread-safe via _kernel_pool_lock; kernel creation happens outside the lock
    since lazy-load may hit Supabase."""
    if not cloud_config.SAAS_MODE:
        return _default_kernel

    key = (scope.user_id or "anon", scope.workbook_id or "default")
    with _kernel_pool_lock:
        if key in _kernel_pool:
            _kernel_pool.move_to_end(key)
            return _kernel_pool[key]

    k = GridOSKernel()
    _register_macros_into_fresh(k)
    # Lazy-load workbook state so any endpoint can be the first touch point —
    # we don't rely on the frontend calling /system/load first (e.g. a kernel
    # that got LRU-evicted must rehydrate silently on the next access).
    if scope.user_id:
        try:
            state = workbook_store.load(scope)
            if state:
                k.apply_state_dict(state)
        except Exception as e:
            print(f"[kernel_pool] lazy-load failed for {key}: {e}")

    with _kernel_pool_lock:
        # A concurrent request may have created one first — dedup by returning
        # theirs so both see the same state.
        if key in _kernel_pool:
            _kernel_pool.move_to_end(key)
            return _kernel_pool[key]
        _kernel_pool[key] = k
        while len(_kernel_pool) > KERNEL_POOL_MAX:
            _kernel_pool.popitem(last=False)
    return k


async def current_kernel_dep(
    user: AuthUser = Depends(require_user),
    x_workbook_id: Optional[str] = Header(None, alias="X-Workbook-Id"),
    workbook_id: Optional[str] = Query(None),
) -> GridOSKernel:
    """FastAPI dep that resolves the request's kernel and binds it to the
    ContextVar for the duration of the request. Every endpoint that reads or
    writes workbook state MUST declare this dep — otherwise the `kernel`
    proxy raises in SaaS mode.

    Must be `async def` — sync deps run in a threadpool with a *copied*
    context, so `_current_kernel.set(k)` would land in the thread's copy and
    the async endpoint body would see an empty ContextVar. Running inline in
    the endpoint's task makes the set visible to subsequent `kernel.X` reads.

    workbook_id resolution order: query param (back-compat with endpoints like
    /system/save that already accept it) → X-Workbook-Id header (default path
    for everything else) → _scope_for fallback (user.id)."""
    wb_id = workbook_id or x_workbook_id
    scope = _scope_for(user, wb_id)
    k = _kernel_for_scope(scope)
    _current_kernel.set(k)
    return k


# ---------- Provider registry + API-key storage ----------

PROVIDER_CLASSES: dict[str, type[Provider]] = {
    "gemini": GeminiProvider,
    "anthropic": AnthropicProvider,
    "groq": GroqProvider,
    "openrouter": OpenRouterProvider,
}
PROVIDER_DISPLAY_NAMES = {
    "gemini": "Google Gemini",
    "anthropic": "Anthropic Claude",
    "groq": "Groq",
    "openrouter": "OpenRouter",
}
PROVIDERS: dict[str, Provider] = {}
API_KEYS: dict[str, str] = {}


def _load_api_keys_from_disk() -> dict[str, str]:
    if not API_KEYS_PATH.exists():
        return {}
    try:
        raw = json.loads(API_KEYS_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    if not isinstance(raw, dict):
        return {}
    return {k: v for k, v in raw.items() if isinstance(k, str) and isinstance(v, str) and v.strip()}


def _persist_api_keys() -> None:
    # Only write after ensuring parent dir exists and data/ is gitignored (it is).
    API_KEYS_PATH.write_text(json.dumps(API_KEYS, indent=2), encoding="utf-8")


def _seed_keys_from_env(keys: dict[str, str]) -> dict[str, str]:
    """Fall back to environment variables for any provider not already present on disk.
    This preserves the old behaviour where GOOGLE_API_KEY came from .env."""
    env_map = {
        "gemini": os.environ.get("GOOGLE_API_KEY"),
        "anthropic": os.environ.get("ANTHROPIC_API_KEY"),
        "groq": os.environ.get("GROQ_API_KEY"),
        "openrouter": os.environ.get("OPENROUTER_API_KEY"),
    }
    for pid, env_val in env_map.items():
        if pid not in keys and env_val:
            keys[pid] = env_val
    return keys


def _rebuild_providers() -> None:
    """Instantiate provider clients for every key currently configured. Failures are
    isolated — a bad key for one provider doesn't block others."""
    PROVIDERS.clear()
    for provider_id, key in API_KEYS.items():
        cls = PROVIDER_CLASSES.get(provider_id)
        if not cls or not key:
            continue
        try:
            PROVIDERS[provider_id] = cls(api_key=key)
        except Exception as e:
            # Keep the key on disk but don't register a broken client.
            print(f"[providers] Failed to init {provider_id}: {e}")


API_KEYS.update(_seed_keys_from_env(_load_api_keys_from_disk()))
_rebuild_providers()


def _configured_provider_ids() -> set[str]:
    return set(PROVIDERS.keys())


def _resolve_model(model_id: Optional[str]) -> tuple[str, Provider]:
    """Pick a model + provider. Falls back to a sensible default if model_id is
    missing or its provider has no key configured."""
    configured = _configured_provider_ids()
    if not configured:
        raise HTTPException(
            status_code=400,
            detail="No LLM provider is configured. Add an API key in Settings.",
        )
    entry = get_model_entry(model_id) if model_id else None
    if entry and entry["provider"] in PROVIDERS:
        return entry["id"], PROVIDERS[entry["provider"]]
    fallback_id = default_model_id(configured)
    if not fallback_id:
        raise HTTPException(status_code=400, detail="No usable model available.")
    fallback_entry = get_model_entry(fallback_id)
    return fallback_entry["id"], PROVIDERS[fallback_entry["provider"]]


# ---------- Telemetry ----------


def _append_telemetry(entry: dict) -> None:
    existing: list = []
    if TELEMETRY_PATH.exists():
        try:
            existing = json.loads(TELEMETRY_PATH.read_text(encoding="utf-8"))
            if not isinstance(existing, list):
                existing = []
        except json.JSONDecodeError:
            existing = []
    existing.append(entry)
    TELEMETRY_PATH.write_text(json.dumps(existing, indent=2), encoding="utf-8")


def call_model(
    agent_id: str,
    *,
    system_instruction: str,
    user_message: str,
    model_id: Optional[str] = None,
    max_attempts: int = 4,
):
    """Route the call through the configured provider for the requested model.
    Retries on transient errors with exponential backoff (~1s, ~2s, ~4s)."""
    model, provider = _resolve_model(model_id)

    last_exc: Optional[Exception] = None
    response = None
    for attempt in range(1, max_attempts + 1):
        try:
            response = provider.generate(
                model=model,
                system_instruction=system_instruction,
                user_message=user_message,
            )
            break
        except Exception as exc:
            last_exc = exc
            if attempt >= max_attempts or not provider.is_transient_error(exc):
                raise
            delay = (2 ** (attempt - 1)) + random.uniform(0, 0.5)
            time.sleep(delay)
    if response is None:
        if last_exc:
            raise last_exc
        raise RuntimeError("call_model exhausted retries with no response")

    _append_telemetry({
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "agent_id": agent_id,
        "provider": provider.id,
        "model": response.model,
        "prompt_token_count": response.prompt_tokens,
        "candidates_token_count": response.candidates_tokens,
        "total_token_count": response.total_tokens,
        "finish_reason": response.finish_reason,
    })
    # SaaS-only: record this call against the authenticated user + workbook
    # bound on the current request via cloud_usage.set_request_context.
    # No-op in OSS mode or for unauthenticated requests. Best-effort.
    cloud_usage.log_call(
        provider=provider.id,
        model=response.model,
        prompt_tokens=response.prompt_tokens or 0,
        completion_tokens=response.candidates_tokens or 0,
        finish_reason=response.finish_reason,
    )
    return response


def _classify_model_error(exc: Exception) -> str:
    """Return 'transient' | 'auth' | 'other' based on the best provider heuristic we can apply."""
    for provider in PROVIDERS.values():
        if provider.is_auth_error(exc):
            return "auth"
        if provider.is_transient_error(exc):
            return "transient"
    return "other"


# ---------- Request models ----------


class ChatRequest(BaseModel):
    prompt: str
    history: List[Dict[str, str]] = []
    scope: str = "sheet"
    selected_cells: List[str] = []
    sheet: Optional[str] = None
    model_id: Optional[str] = None


class ChainRequest(ChatRequest):
    max_iterations: int = MAX_CHAIN_ITERATIONS


class ApiKeySaveRequest(BaseModel):
    provider: str
    api_key: str


class FormulaRequest(BaseModel):
    function_name: str
    arguments: list[float]


class CellUpdateRequest(BaseModel):
    cell: str
    value: Optional[str] = ""
    sheet: Optional[str] = None


class RangeUpdateRequest(BaseModel):
    target_cell: str
    values: list[list[str | int | float | bool | None]]
    sheet: Optional[str] = None


class PreviewApplyRequest(BaseModel):
    sheet: Optional[str] = None
    agent_id: str
    target_cell: str
    values: list[list]
    shift_direction: str = "right"
    chart_spec: Optional[Dict[str, Any]] = None


class SheetCreateRequest(BaseModel):
    name: Optional[str] = None


class SheetRenameRequest(BaseModel):
    old_name: str
    new_name: str


class SheetActivateRequest(BaseModel):
    name: str


class WorkbookRenameRequest(BaseModel):
    name: str


# ---------- Agent routing & prompts ----------


BASE_SYSTEM_RULES = (
    "You are operating within GridOS. Check the \"locked\" metadata for every cell "
    "before proposing a write. Do not attempt to overwrite locked cells."
)

OUTPUT_FORMAT_SPEC = """
OUTPUT FORMAT: strictly valid JSON (no markdown fences):
{
    "reasoning": "Short explanation.",
    "target_cell": "A1",
    "values": [["..."]],
    "chart_spec": null,
    "macro_spec": null,
    "plan": null
}

If the user is asking for a MULTI-SECTION build (financial model, forecast, 3-statement, operating model, DCF, budget, etc.), emit a `plan` object on the FIRST turn only. It is informational — the system shows it to the user so they can see the full structure before cells are filled:
{
    "title": "Quarterly Operating Model",
    "anchor": "B2",
    "sections": [
        {"label": "Header row",          "target": "B2:F2", "notes": "Metric, Q1..Q4"},
        {"label": "Revenue",             "target": "B3:F3", "notes": "10% QoQ growth from 100"},
        {"label": "COGS",                "target": "B4:F4", "notes": "=MULTIPLY(revenue, 0.4)"}
    ]
}
On this first turn, `values` contains ONLY the first section. On subsequent chain turns, omit `plan` (set null) and write the NEXT section. Emit values=[[""]] when every section is done.

If the user asks for a chart/graph/visualization, also fill in chart_spec:
{
    "data_range": "A1:B6",        // rectangular range covering labels + values
    "chart_type": "bar",          // one of: bar, line, pie
    "title": "Scores",
    "anchor_cell": "D2",          // top-left cell where the chart overlay appears; pick an empty area
    "orientation": "columns"      // "columns" = first column is labels (typical); "rows" = first row is labels
}
Omit chart_spec (leave as null) when the user is only writing data or editing cells.

If the user wants to MOVE, RESIZE, RETYPE, or RENAME an EXISTING chart (e.g. "place the performance chart at F16", "make the scores chart a pie"), reuse the SAME title as the existing chart in chart_spec — the system will update it in place. In that case set "values": null and set "target_cell" to the new anchor_cell. Do NOT invent placeholder data.

If the user asks for a NEW named formula/metric that is not already listed in USER MACROS or the built-in primitives, you MAY propose a new macro by filling in macro_spec:
{
    "name": "MARGIN",                    // unique identifier, letters/digits/underscore only
    "params": ["A", "B"],                // parameter names (uppercase letters)
    "description": "Gross margin: (A - B) / A",
    "body": "=DIVIDE(MINUS(A, B), A)"   // MUST only call registered primitives. Nested primitive calls ARE allowed here (this is the one place nesting is permitted — macro bodies are composed expressions). No infix operators, no references to other user macros.
}
Proposed macros are NOT saved automatically — the user reviews and approves them. In the SAME response, do NOT write any cell values that call the proposed macro (it isn't registered yet). Keep "values" null or write unrelated cells. The user will re-ask after approval to use the new macro.
""".strip()


# Router prefers the fastest configured small model — classification is trivial
# and doesn't need frontier quality. Wall-clock savings are visible on every chat.
# Ordered fastest-first; first entry whose provider has a key wins.
_ROUTER_MODEL_PREFERENCE = [
    ("openai/gpt-oss-20b", "groq"),              # ~1000 tps
    ("llama-3.1-8b-instant", "groq"),            # ~560 tps
    ("gemini-3.1-flash-lite-preview", "gemini"), # Google's fastest
    ("claude-haiku-4-5-20251001", "anthropic"),  # Anthropic's fastest
]


def _pick_router_model(user_choice: Optional[str]) -> Optional[str]:
    configured = _configured_provider_ids()
    for mid, pid in _ROUTER_MODEL_PREFERENCE:
        if pid in configured:
            return mid
    return user_choice


def route_prompt(prompt: str, history_context: str, model_id: Optional[str] = None) -> str:
    if len(AGENTS) == 1:
        return next(iter(AGENTS))

    options = "\n".join(
        f"- {agent['id']}: {agent.get('router_description', agent.get('display_name', agent['id']))}"
        for agent in AGENTS.values()
    )
    instruction = f"""
Analyze this user task: "{prompt}".
Previous context: {history_context}

Available agent profiles:
{options}

Return ONLY the lowercase agent id that best fits the task. No other text.
""".strip()

    res = call_model(
        "router",
        system_instruction="You are a routing classifier. Respond with only a lowercase agent id.",
        user_message=instruction,
        model_id=_pick_router_model(model_id),
    )
    candidate = res.text.strip().lower().split()[0] if res.text else "general"
    return candidate if candidate in AGENTS else "general"


def build_system_instruction(agent: dict, context: dict, req: ChatRequest) -> str:
    selected_summary = ", ".join(req.selected_cells) if req.selected_cells else "No cells selected."
    scope_line = "Selected cells only" if req.scope == "selection" else "Entire active sheet"
    bounds = context.get("occupied_bounds")
    bounds_line = (
        f"Occupied region: {bounds['top_left']} -> {bounds['bottom_right']} "
        f"({bounds['rows']} rows x {bounds['cols']} cols)"
        if bounds else "Occupied region: empty"
    )

    existing_charts = kernel.list_charts(req.sheet or kernel.active_sheet)
    if existing_charts:
        chart_lines = [
            f"- \"{c.get('title') or '(untitled)'}\" ({c.get('chart_type')}, range {c.get('data_range')}, anchor {c.get('anchor_cell')})"
            for c in existing_charts
        ]
        charts_section = "EXISTING CHARTS ON THIS SHEET (reuse the same title in chart_spec to update one):\n" + "\n".join(chart_lines)
    else:
        charts_section = "EXISTING CHARTS ON THIS SHEET: none"

    if USER_MACROS:
        macro_lines = [
            f"- {m['name']}({', '.join(m['params'])}) — {m.get('description') or 'user macro'}"
            for m in USER_MACROS
        ]
        macros_section = (
            "USER MACROS (callable like built-in formulas, single flat call, no nesting in the grid):\n"
            + "\n".join(macro_lines)
        )
    else:
        macros_section = "USER MACROS: none"

    enabled_hero_ids = [t["id"] for t in HERO_TOOLS_CATALOG if HERO_TOOLS_STATE.get(t["id"])]
    if enabled_hero_ids:
        hero_lines = [
            f"- {t['display_name']}: {t['description']}"
            for t in HERO_TOOLS_CATALOG
            if t["id"] in enabled_hero_ids
        ]
        hero_section = (
            "HERO TOOLS ENABLED (advisory — they are not wired up yet, mention capability only if the user asks):\n"
            + "\n".join(hero_lines)
        )
    else:
        hero_section = "HERO TOOLS ENABLED: none"

    primitive_names = _builtin_primitive_names()
    primitives_section = (
        "AVAILABLE PRIMITIVES (authoritative — if a function is not in this list, it does NOT exist; "
        "do not invent names like POWER, LN, IF, etc. unless they appear here):\n"
        + ", ".join(primitive_names)
    )

    sections = [
        BASE_SYSTEM_RULES,
        f"ACTIVE SHEET: {req.sheet or kernel.active_sheet}\nVIEW SCOPE: {scope_line}\nSELECTED CELLS: {selected_summary}\n{bounds_line}",
        f"CELL METADATA (a1 -> {{val, locked, type}}):\n{context['cell_metadata_json']}",
        f"READABLE GRID STATE:\n{context['formatted_data']}",
        charts_section,
        primitives_section,
        macros_section,
        hero_section,
    ]

    if req.history:
        history_lines = "\n".join(f"{h['role'].upper()}: {h['content']}" for h in req.history)
        sections.append(
            "CONVERSATION HISTORY (oldest first — the first user message is the original task; "
            "check it for any targets you have not yet written):\n" + history_lines
        )

    sections.extend([agent["system_prompt"], OUTPUT_FORMAT_SPEC])
    return "\n\n".join(sections)


def _extract_first_json_object(text: str) -> Optional[str]:
    """Find the first balanced {...} in text, tolerant of prose prefix/suffix.
    Handles quoted strings + escape chars so braces inside a JSON string don't
    confuse the depth counter. Returns None if no complete object is present."""
    start = text.find("{")
    if start == -1:
        return None
    depth = 0
    in_string = False
    escape = False
    for i in range(start, len(text)):
        ch = text[i]
        if escape:
            escape = False
            continue
        if in_string:
            if ch == "\\":
                escape = True
            elif ch == '"':
                in_string = False
            continue
        if ch == '"':
            in_string = True
        elif ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return text[start : i + 1]
    return None


def _parse_ai_response(response) -> dict:
    """Extract JSON payload from a ProviderResponse. Raises HTTPException(422) with
    a user-actionable message when the model returned empty/malformed output —
    this is the common Groq/OpenRouter failure mode on complex prompts."""
    text = (response.text or "").replace("```json", "").replace("```", "").strip()

    finish = response.finish_reason
    ctx = f"{response.provider_id}/{response.model}"
    if finish:
        ctx += f" (finish_reason={finish})"

    if not text:
        hint = (
            "hit the output-token cap — try a shorter prompt or a model with more headroom"
            if finish and ("length" in str(finish).lower() or str(finish).upper() == "MAX_TOKENS")
            else "returned no content — try a stronger model (Gemini/Claude) or rephrase"
        )
        raise HTTPException(
            status_code=422,
            detail=f"Model {ctx} {hint}.",
        )

    # Try a strict parse first; if that fails, dig out the first balanced {...} —
    # small open models often prepend prose like "Sure! Here's the JSON: {...}".
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    extracted = _extract_first_json_object(text)
    if extracted:
        try:
            return json.loads(extracted)
        except json.JSONDecodeError:
            pass

    preview = text[:180].replace("\n", " ")
    raise HTTPException(
        status_code=422,
        detail=(
            f"Model {ctx} returned non-JSON output — try a stronger model (Gemini/Claude) "
            f"or rephrase. First bytes: {preview!r}"
        ),
    )


def _sanitize_plan(raw: Any) -> Optional[dict]:
    """Coerce an agent-emitted plan into a shape the UI can render safely. Returns None if empty."""
    if not isinstance(raw, dict):
        return None
    sections_raw = raw.get("sections")
    if not isinstance(sections_raw, list) or not sections_raw:
        return None
    sections = []
    for item in sections_raw:
        if not isinstance(item, dict):
            continue
        label = str(item.get("label") or "").strip()
        target = str(item.get("target") or "").strip()
        notes = str(item.get("notes") or "").strip()
        if not label and not target and not notes:
            continue
        sections.append({"label": label, "target": target, "notes": notes})
    if not sections:
        return None
    return {
        "title": str(raw.get("title") or "").strip(),
        "anchor": str(raw.get("anchor") or "").strip(),
        "sections": sections,
    }


def _validate_proposed_macro(raw: Any) -> tuple[Optional[dict], Optional[str]]:
    """Dry-compile an agent-proposed macro. Returns (normalized_spec, error_message)."""
    if not raw or not isinstance(raw, dict):
        return None, None
    name = str(raw.get("name") or "").strip()
    body = str(raw.get("body") or "").strip()
    if not name or not body:
        return None, "Macro proposal missing name or body."
    params_raw = raw.get("params") or []
    if not isinstance(params_raw, list):
        return None, "Macro params must be a list."
    params = [str(p).strip() for p in params_raw if str(p).strip()]

    # Exclude the same name from the registry ONLY if it's currently a user macro
    # (so proposing an update to an existing macro is allowed). Built-in primitives
    # must still collision-check, because macros can't shadow them.
    upper = name.upper()
    existing_macro_names = _macro_names()
    if upper in existing_macro_names:
        registry = {k: v for k, v in kernel.evaluator.registry.items() if k.upper() != upper}
    else:
        registry = dict(kernel.evaluator.registry)
    try:
        compile_macro(name=name, params=params, body=body, registry=registry)
    except MacroError as e:
        return None, str(e)

    return {
        "name": upper,
        "params": [p.upper() for p in params],
        "description": str(raw.get("description") or "").strip(),
        "body": body,
        "replaces_existing": upper in _macro_names(),
    }, None


def generate_agent_preview(req: ChatRequest) -> dict:
    sheet = req.sheet or kernel.active_sheet
    context = kernel.get_context_for_ai(sheet, req.selected_cells, req.scope)
    history_context = "\n".join([f"{h['role'].upper()}: {h['content']}" for h in req.history])

    agent_id = route_prompt(req.prompt, history_context, model_id=req.model_id)
    agent = AGENTS[agent_id]
    system_instruction = build_system_instruction(agent, context, req)

    final_response = call_model(
        agent_id,
        system_instruction=system_instruction,
        user_message=req.prompt,
        model_id=req.model_id,
    )
    ai_data = _parse_ai_response(final_response)

    raw_values = ai_data.get("values")
    raw_target = ai_data.get("target_cell")
    chart_spec = ai_data.get("chart_spec")
    proposed_macro, macro_error = _validate_proposed_macro(ai_data.get("macro_spec"))
    plan = _sanitize_plan(ai_data.get("plan"))
    fallback_target = req.selected_cells[0] if req.selected_cells else "A1"

    has_values = isinstance(raw_values, list) and any(
        any(v not in ("", None) for v in row) for row in raw_values if isinstance(row, list)
    )

    if not has_values:
        # No grid write this turn. Covers three cases cleanly:
        #   1. Pure acknowledgement ("hello" → the agent has no cells to write).
        #   2. Chart-only / macro-proposal-only turn.
        #   3. Plan-declaration turn that deferred the first write.
        # The UI hides the Apply button when values is null and no chart is attached.
        return {
            "category": agent_id,
            "reasoning": ai_data.get("reasoning"),
            "sheet": sheet,
            "scope": req.scope,
            "selected_cells": req.selected_cells,
            "agent_id": agent_id,
            "target_cell": raw_target or fallback_target,
            "original_request": raw_target or fallback_target,
            "preview_cells": [],
            "values": None,
            "chart_spec": chart_spec,
            "proposed_macro": proposed_macro,
            "macro_error": macro_error,
            "plan": plan,
        }

    intent = AgentIntent(
        agent_id=agent_id,
        target_start_a1=raw_target or fallback_target,
        data_payload=raw_values,
        shift_direction="right",
    )

    preview = kernel.preview_agent_intent(intent, sheet)

    dep_issues = _find_empty_formula_deps(
        preview["preview_cells"],
        kernel._sheet_state(sheet),
    )
    if dep_issues:
        bullets = "\n".join(
            f"  - {d['cell']} ({d['formula']}) references empty cell(s): {', '.join(d['empty_refs'])}"
            for d in dep_issues[:5]
        )
        raise HTTPException(
            status_code=422,
            detail=(
                "The agent proposed formulas whose inputs are empty — applying would produce "
                "#DIV/0! or misleading zeros. Re-ask the agent to also populate the referenced "
                "cells, or fill them yourself first.\n" + bullets
            ),
        )

    return {
        "category": agent_id,
        "reasoning": ai_data.get("reasoning"),
        "sheet": sheet,
        "scope": req.scope,
        "selected_cells": req.selected_cells,
        "agent_id": agent_id,
        "target_cell": preview["actual_target"],
        "original_request": preview["original_target"],
        "preview_cells": preview["preview_cells"],
        "values": raw_values,
        "chart_spec": chart_spec,
        "proposed_macro": proposed_macro,
        "macro_error": macro_error,
        "plan": plan,
    }


# ---------- Endpoints ----------


@app.get("/healthz")
async def healthz():
    """Liveness probe for Render (and UptimeRobot keep-warm). Returns 200
    unconditionally as long as the ASGI app is importable and serving —
    no DB or external-provider touch, because those are fine to be down
    while the process itself is alive."""
    return {"ok": True}


@app.get("/")
async def serve_landing():
    return FileResponse("static/landing.html")


@app.get("/workbook")
async def serve_workbook():
    return FileResponse("static/index.html")


@app.get("/login")
async def serve_login():
    """Login page — only meaningful in SaaS mode. In OSS mode the page's own
    bootstrap redirects to / since /cloud/status reports mode=oss."""
    return FileResponse("static/login.html")


@app.get("/agents")
async def list_agents():
    return {
        "agents": [
            {"id": a["id"], "display_name": a.get("display_name", a["id"]), "router_description": a.get("router_description", "")}
            for a in AGENTS.values()
        ]
    }


@app.post("/agent/chat")
async def chat_with_agent(
    req: ChatRequest,
    user: AuthUser = Depends(require_user),
    k: GridOSKernel = Depends(current_kernel_dep),
):
    # Bind the request's user + workbook scope so every call_model() downstream
    # records usage against this user. In OSS mode user.id == "oss" and
    # cloud_usage.log_call is a no-op.
    # workbook_id is intentionally None — tagging by workbook requires a row in
    # public.workbooks, which only exists after the user's first /system/save.
    # FK would reject the usage_logs insert otherwise. Revisit once multi-workbook
    # UX lands and workbook rows are provisioned eagerly.
    cloud_usage.set_request_context(user.id, None)
    if cloud_config.SAAS_MODE:
        try:
            cloud_usage.over_quota_check(user.id)
        except cloud_usage.QuotaExceeded as qe:
            raise HTTPException(status_code=402, detail={
                "message": "Monthly token cap reached for your tier.",
                "usage": qe.summary,
            })
    try:
        return generate_agent_preview(req)
    except HTTPException:
        raise
    except Exception as e:
        kind = _classify_model_error(e)
        if kind == "transient":
            raise HTTPException(
                status_code=503,
                detail="Model provider is temporarily overloaded (tried 4x with backoff). Wait a moment and try again.",
            )
        if kind == "auth":
            raise HTTPException(
                status_code=401,
                detail="The API key for this model was rejected. Update it in Settings.",
            )
        raise HTTPException(status_code=500, detail=f"Agent Error: {str(e)}")


@app.post("/agent/apply")
async def apply_agent_preview(
    req: PreviewApplyRequest,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    intent = AgentIntent(
        agent_id=req.agent_id,
        target_start_a1=req.target_cell,
        data_payload=req.values,
        shift_direction=req.shift_direction,
    )
    requested, actual = kernel.process_agent_intent(intent, req.sheet)
    chart = None
    if req.chart_spec:
        try:
            chart = kernel.add_chart(req.chart_spec, sheet_name=req.sheet)
        except Exception as e:
            return {
                "status": "Partial",
                "sheet": req.sheet or kernel.active_sheet,
                "actual_target": actual,
                "chart_error": f"Chart skipped: {e}",
            }
    return {
        "status": "Success" if requested == actual else "Collision Resolved",
        "sheet": req.sheet or kernel.active_sheet,
        "actual_target": actual,
        "chart": chart,
    }


_CELL_REF_RE = re.compile(r"[A-Z]+\d+")


def _find_empty_formula_deps(preview_cells: list[dict], sheet_state: dict) -> list[dict]:
    """For each formula-bearing preview cell, flag any cell reference that points
    at an empty cell in the current sheet AND isn't being populated by this same
    preview. Catches the '#DIV/0! from an empty baseline' bug class — e.g. the
    agent writes =GROWTH(C4, C3) but forgets to seed C3."""
    self_written_nonempty: set[str] = set()
    for p in preview_cells:
        v = p.get("value")
        if v not in (None, ""):
            self_written_nonempty.add(p["cell"].upper())

    issues: list[dict] = []
    for p in preview_cells:
        v = p.get("value")
        if not isinstance(v, str) or not v.startswith("="):
            continue
        empty_refs: list[str] = []
        for ref in _CELL_REF_RE.findall(v.upper()):
            if ref in self_written_nonempty:
                continue
            try:
                r, c = a1_to_coords(ref)
            except ValueError:
                continue
            cell = sheet_state["cells"].get((r, c))
            if cell is None:
                empty_refs.append(ref)
            elif cell.value in (None, "") and not cell.formula:
                empty_refs.append(ref)
        if empty_refs:
            issues.append({
                "cell": p["cell"],
                "formula": v,
                "empty_refs": empty_refs,
            })
    return issues


def _formula_references_text_cell(formula: str, sheet_state: dict) -> list[str]:
    """Scan cell references inside a formula; return the A1 names of any that
    resolve to a non-numeric value (typical off-by-one symptom: formula pointing
    at a row-label column). Empty list means all refs look numeric."""
    bad_refs: list[str] = []
    for ref in _CELL_REF_RE.findall(formula.upper()):
        try:
            r, c = a1_to_coords(ref)
        except ValueError:
            continue
        cell = sheet_state["cells"].get((r, c))
        if cell is None:
            continue  # empty cell is fine (coerced to 0, intentional)
        v = cell.value
        if isinstance(v, bool):
            continue
        if isinstance(v, (int, float)):
            continue
        if v in ("", None):
            continue
        bad_refs.append(ref)
    return bad_refs


def _observe_written_cells(preview_cells: list[dict], sheet: str) -> list[dict]:
    """After applying, collect computed values for each just-written cell. Attach a
    'warning' field when the formula dereferenced a text-valued cell — the classic
    off-by-one-column mistake the agent makes on labeled rows."""
    state = kernel._sheet_state(sheet)
    observations = []
    for item in preview_cells:
        a1 = item["cell"]
        try:
            r, c = a1_to_coords(a1)
        except ValueError:
            continue
        cell = state["cells"].get((r, c))
        if cell is None:
            continue
        warning = None
        if cell.formula:
            bad_refs = _formula_references_text_cell(cell.formula, state)
            if bad_refs:
                labels = ", ".join(f"{ref}={state['cells'][a1_to_coords(ref)].value!r}" for ref in bad_refs)
                warning = (
                    f"Formula in {a1} references non-numeric cell(s): {labels}. "
                    f"This is the COLUMN ALIGNMENT bug — the formula in column "
                    f"{a1.rstrip('0123456789')} must reference cells in column "
                    f"{a1.rstrip('0123456789')}, not a label column."
                )
        observations.append({
            "cell": a1,
            "value": cell.value,
            "formula": cell.formula,
            "warning": warning,
        })
    return observations


def _is_completion_signal(values) -> bool:
    """An agent signals 'task complete' by returning an empty-string grid."""
    if not values:
        return True
    for row in values:
        for v in row:
            if v not in ("", None):
                return False
    return True


@app.post("/agent/chat/chain")
async def chat_chain(
    req: ChainRequest,
    user: AuthUser = Depends(require_user),
    k: GridOSKernel = Depends(current_kernel_dep),
):
    """Auto-apply the agent's writes, observe formula results, and loop up to max_iterations times."""
    # workbook_id is intentionally None — tagging by workbook requires a row in
    # public.workbooks, which only exists after the user's first /system/save.
    # FK would reject the usage_logs insert otherwise. Revisit once multi-workbook
    # UX lands and workbook rows are provisioned eagerly.
    cloud_usage.set_request_context(user.id, None)
    if cloud_config.SAAS_MODE:
        try:
            cloud_usage.over_quota_check(user.id)
        except cloud_usage.QuotaExceeded as qe:
            raise HTTPException(status_code=402, detail={
                "message": "Monthly token cap reached for your tier.",
                "usage": qe.summary,
            })
    try:
        sheet = req.sheet or kernel.active_sheet
        history = list(req.history)
        steps: list[dict] = []
        current_prompt = req.prompt
        max_iters = max(1, min(req.max_iterations, MAX_CHAIN_ITERATIONS))

        # Plan-level state persists across iterations: the agent declares its plan
        # once on turn 1, then subsequent turns omit it. We need to remember it so
        # the chain keeps running until every section is written (or max_iters).
        active_plan: Optional[dict] = None

        for iteration in range(max_iters):
            chat_req = ChatRequest(
                prompt=current_prompt,
                history=history,
                scope=req.scope,
                selected_cells=req.selected_cells,
                sheet=sheet,
                model_id=req.model_id,
            )
            preview = generate_agent_preview(chat_req)
            values = preview["values"] or [[""]]
            chart_spec = preview.get("chart_spec")
            proposed_macro = preview.get("proposed_macro")
            macro_error = preview.get("macro_error")
            plan = preview.get("plan")
            if plan and active_plan is None:
                active_plan = plan
            skip_cell_write = preview["values"] is None and (chart_spec is not None or proposed_macro is not None)

            if _is_completion_signal(values) and not skip_cell_write:
                steps.append({
                    "iteration": iteration,
                    "agent_id": preview["agent_id"],
                    "reasoning": preview["reasoning"],
                    "target": preview["original_request"],
                    "values": values,
                    "observations": [],
                    "completion_signal": True,
                    "proposed_macro": proposed_macro,
                    "macro_error": macro_error,
                    "plan": plan,
                })
                break

            if skip_cell_write:
                actual_target = preview["original_request"]
                observations = []
                formula_observations = []
            else:
                intent = AgentIntent(
                    agent_id=preview["agent_id"],
                    target_start_a1=preview["original_request"],
                    data_payload=values,
                    shift_direction="right",
                )
                _, actual_target = kernel.process_agent_intent(intent, sheet)
                observations = _observe_written_cells(preview["preview_cells"], sheet)
                formula_observations = [o for o in observations if o["formula"]]

            chart = None
            chart_error = None
            if chart_spec:
                try:
                    chart = kernel.add_chart(chart_spec, sheet_name=sheet)
                except Exception as e:
                    chart_error = str(e)

            steps.append({
                "iteration": iteration,
                "agent_id": preview["agent_id"],
                "reasoning": preview["reasoning"],
                "target": actual_target,
                "values": values,
                "observations": observations,
                "completion_signal": False,
                "chart": chart,
                "chart_error": chart_error,
                "proposed_macro": proposed_macro,
                "macro_error": macro_error,
                "plan": plan,
            })

            assistant_payload = {
                "reasoning": preview["reasoning"],
                "target": actual_target,
                "values": values,
            }
            if plan:
                assistant_payload["plan"] = plan

            history.append({"role": "user", "content": current_prompt})
            history.append({"role": "assistant", "content": json.dumps(assistant_payload)})

            # Continue the chain while an active plan still has sections left, or while we're
            # seeing formulas (legacy follow-up heuristic). Break only when neither applies.
            # Only count a section as "written" if the step had no column-alignment warnings —
            # otherwise the agent needs to retry the SAME section, not advance.
            non_completion_steps = [s for s in steps if not s.get("completion_signal")]
            sections_written = sum(
                1
                for s in non_completion_steps
                if not any(o.get("warning") for o in s.get("observations", []))
            )
            plan_sections_total = len(active_plan.get("sections", [])) if active_plan else 0
            plan_remaining = plan_sections_total - sections_written
            has_plan_work = plan_remaining > 0
            has_retry_work = any(o.get("warning") for o in observations)
            if not formula_observations and not has_plan_work and not has_retry_work:
                break

            obs_part = ""
            if formula_observations:
                summary = ", ".join(f"{o['cell']}={o['value']}" for o in formula_observations)
                obs_part = f"Observed after last write: [{summary}]. "

            warning_obs = [o for o in observations if o.get("warning")]
            warning_part = ""
            if warning_obs:
                bullets = "\n".join(f"- {o['warning']}" for o in warning_obs)
                warning_part = (
                    "\n\n*** COLUMN ALIGNMENT WARNINGS — YOU MUST FIX THESE BEFORE MOVING ON ***\n"
                    f"{bullets}\n"
                    "Re-emit the SAME section (same target cell) with corrected formulas. "
                    "Each formula in column X must reference cells in column X, not a label column. "
                    "Do NOT move to the next section until the current section has no warnings.\n\n"
                )

            plan_part = ""
            if has_plan_work:
                next_section = active_plan["sections"][sections_written] if sections_written < plan_sections_total else None
                next_hint = ""
                if next_section:
                    bits = []
                    if next_section.get("label"):
                        bits.append(f"label \"{next_section['label']}\"")
                    if next_section.get("target"):
                        bits.append(f"target range {next_section['target']}")
                    if next_section.get("notes"):
                        bits.append(f"notes: {next_section['notes']}")
                    if bits:
                        label = "The section to (re-)write is" if has_retry_work else "The next section is"
                        next_hint = f" {label}: " + "; ".join(bits) + "."
                verb = "re-emit" if has_retry_work else "write"
                plan_part = (
                    f"You declared a {plan_sections_total}-section plan on turn 1. "
                    f"{sections_written} section(s) have been written cleanly so far. "
                    f"{plan_remaining} section(s) remain (or need retry).{next_hint} "
                    f"{verb.capitalize()} ONLY that one section now. "
                    "Do NOT re-emit the plan (set plan=null). "
                    "When every section is done, signal completion by returning values=[[\"\"]]."
                )

            current_prompt = (
                warning_part
                + obs_part
                + plan_part
                + " If the ORIGINAL task has more targets left to write, produce the next one now and do NOT "
                "repeat cells you have already written. If every part of the original task is done, signal "
                "completion by returning values=[[\"\"]] with target_cell equal to the last written cell."
            )

        return {
            "sheet": sheet,
            "steps": steps,
            "iterations_used": len(steps),
            "terminated_early": len(steps) < max_iters,
        }
    except HTTPException:
        raise
    except Exception as e:
        kind = _classify_model_error(e)
        if kind == "transient":
            raise HTTPException(
                status_code=503,
                detail="Model provider is temporarily overloaded (tried 4x with backoff). Wait a moment and try again.",
            )
        if kind == "auth":
            raise HTTPException(
                status_code=401,
                detail="The API key for this model was rejected. Update it in Settings.",
            )
        raise HTTPException(status_code=500, detail=f"Chain Error: {str(e)}")


@app.post("/agent/write", response_model=WriteResponse)
async def agent_write(
    intent: AgentIntent,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    try:
        requested_a1, actual_a1 = kernel.process_agent_intent(intent)
        return WriteResponse(
            status="Success" if requested_a1 == actual_a1 else "Collision Resolved",
            original_target=requested_a1,
            actual_target=actual_a1,
            message=f"Wrote data starting at {actual_a1}",
        )
    except Exception as e:
        return WriteResponse(
            status="Error",
            original_target=intent.target_start_a1,
            actual_target=intent.target_start_a1,
            message=str(e),
        )


@app.get("/debug/grid")
async def get_grid(
    sheet: Optional[str] = None,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    target = sheet or kernel.active_sheet
    return {
        "sheet": target,
        "cells": kernel.export_sheet(target),
        "charts": kernel.list_charts(target),
    }


@app.get("/api/workbook")
async def get_workbook(k: GridOSKernel = Depends(current_kernel_dep)):
    return {
        "workbook_name": kernel.workbook_name,
        "active_sheet": kernel.active_sheet,
        "sheets": kernel.list_sheets(),
        "chat_log": list(kernel.chat_log),
    }


class ChatLogReplaceRequest(BaseModel):
    entries: List[Dict[str, Any]] = []


@app.post("/workbook/chat/replace")
async def replace_chat_log(
    req: ChatLogReplaceRequest,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    try:
        kernel.set_chat_log(req.entries)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"status": "Success", "count": len(kernel.chat_log)}


@app.post("/workbook/chat/clear")
async def clear_chat_log(k: GridOSKernel = Depends(current_kernel_dep)):
    kernel.clear_chat_log()
    return {"status": "Success"}


@app.post("/workbook/rename")
async def rename_workbook(
    req: WorkbookRenameRequest,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    try:
        name = kernel.rename_workbook(req.name)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"workbook_name": name}


@app.post("/workbook/sheet")
async def create_sheet(
    req: SheetCreateRequest,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    name = kernel.create_sheet(req.name)
    return {"sheet": name, "sheets": kernel.list_sheets(), "active_sheet": kernel.active_sheet}


@app.post("/workbook/sheet/rename")
async def rename_sheet(
    req: SheetRenameRequest,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    name = kernel.rename_sheet(req.old_name, req.new_name)
    return {"sheet": name, "sheets": kernel.list_sheets(), "active_sheet": kernel.active_sheet}


@app.post("/workbook/sheet/activate")
async def activate_sheet(
    req: SheetActivateRequest,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    name = kernel.activate_sheet(req.name)
    return {"sheet": name, "sheets": kernel.list_sheets(), "active_sheet": kernel.active_sheet}


@app.post("/grid/cell")
async def update_cell(
    req: CellUpdateRequest,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    try:
        target = kernel.write_user_cell(req.cell.upper(), req.value, user_id="User", sheet_name=req.sheet)
        return {"status": "Success", "cell": target, "sheet": req.sheet or kernel.active_sheet}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/grid/range")
async def update_range(
    req: RangeUpdateRequest,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    try:
        target = kernel.write_user_range(req.target_cell.upper(), req.values, user_id="User", sheet_name=req.sheet)
        return {"status": "Success", "target": target, "sheet": req.sheet or kernel.active_sheet}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/auth/whoami")
async def whoami(user: AuthUser = Depends(require_user)):
    """Return the signed-in user's claims so the frontend can confirm the
    session is valid before rendering protected UI."""
    return {"id": user.id, "email": user.email, "mode": "saas" if cloud_config.SAAS_MODE else "oss"}


@app.get("/usage/me")
async def usage_me(user: AuthUser = Depends(require_user)):
    """Account summary for the signed-in user: tier + this month's token usage
    + cost estimate. Reads `public.users` (tier, created_at) and
    `public.user_usage` (month rollup). Returns zeros rather than 404 for
    brand-new users who haven't made any calls yet."""
    if not cloud_config.SAAS_MODE:
        # OSS has no authenticated user; return a minimal stub so the same UI
        # code can render something meaningful in local mode.
        return {
            "mode": "oss",
            "email": None,
            "tier": "oss",
            "joined_at": None,
            "month": datetime.now(timezone.utc).strftime("%Y-%m-01"),
            "total_tokens": 0,
            "cost_cents": 0,
            "tier_limit": 0,
            "tokens_remaining": None,
            "quota_pct": 0,
        }
    if not cloud_config.SAAS_FEATURES["usage_tracking"].enabled:
        raise HTTPException(status_code=503, detail="Usage tracking is not configured on this deployment.")
    try:
        from supabase import create_client  # type: ignore
    except ImportError as e:
        raise HTTPException(status_code=503, detail="supabase-py is not installed.") from e

    client = create_client(cloud_config.SUPABASE_URL, cloud_config.SUPABASE_SERVICE_ROLE_KEY)
    month_str = datetime.now(timezone.utc).strftime("%Y-%m-01")

    tier = "free"
    joined_at = None
    try:
        u = (
            client.table("users")
            .select("subscription_tier, created_at")
            .eq("id", user.id)
            .limit(1)
            .execute()
        )
        if u.data:
            tier = u.data[0].get("subscription_tier") or "free"
            joined_at = u.data[0].get("created_at")
    except Exception:
        # Row may not yet exist if the on_auth_user_created trigger lagged;
        # fall through with defaults.
        pass

    total_tokens = 0
    cost_cents = 0
    try:
        usage = (
            client.table("user_usage")
            .select("total_tokens, cost_cents")
            .eq("user_id", user.id)
            .eq("month", month_str)
            .limit(1)
            .execute()
        )
        if usage.data:
            total_tokens = int(usage.data[0].get("total_tokens") or 0)
            cost_cents = int(usage.data[0].get("cost_cents") or 0)
    except Exception:
        pass

    limit = cloud_config.tier_limit(tier)
    if limit > 0:
        remaining = max(0, limit - total_tokens)
        pct = min(100, int(round((total_tokens / limit) * 100))) if limit else 0
    else:
        remaining = None  # unlimited
        pct = 0
    return {
        "mode": "saas",
        "email": user.email,
        "tier": tier,
        "joined_at": joined_at,
        "month": month_str,
        "total_tokens": total_tokens,
        "cost_cents": cost_cents,
        "tier_limit": limit,
        "tokens_remaining": remaining,
        "quota_pct": pct,
    }


@app.post("/system/save")
async def save_grid(
    workbook_id: Optional[str] = None,
    user: AuthUser = Depends(require_user),
    k: GridOSKernel = Depends(current_kernel_dep),
):
    scope = _scope_for(user, workbook_id)
    workbook_store.save(scope, kernel.export_state_dict())
    return {"status": "Success", "workbook_id": scope.workbook_id}


@app.post("/system/load")
async def load_grid(
    workbook_id: Optional[str] = None,
    user: AuthUser = Depends(require_user),
    k: GridOSKernel = Depends(current_kernel_dep),
):
    scope = _scope_for(user, workbook_id)
    state = workbook_store.load(scope)
    if state is None:
        return {"status": "Error", "message": "No save file found."}
    kernel.apply_state_dict(state)
    return {"status": "Success", "workbook_id": scope.workbook_id}


# ---- Multi-workbook endpoints (SaaS) ---------------------------------------
# List / create / delete / rename. All auth-gated; in OSS mode they return a
# 404 so the frontend can cleanly treat them as "feature not available."


class WorkbookCreateRequest(BaseModel):
    title: Optional[str] = None


class WorkbookRenameRequest(BaseModel):
    title: str


def _require_saas_storage() -> None:
    """Guard: endpoints that only make sense in SaaS with a Supabase store."""
    if not cloud_config.SAAS_MODE:
        raise HTTPException(status_code=404, detail="Multi-workbook is a SaaS feature.")
    if not cloud_config.SAAS_FEATURES["cloud_storage"].enabled:
        raise HTTPException(status_code=503, detail="Cloud storage is not configured.")


@app.get("/workbooks")
async def list_workbooks(user: AuthUser = Depends(require_user)):
    """Return the signed-in user's workbooks ordered most-recently-used first,
    plus slot usage so the UI can render the quota badge without a second
    round-trip."""
    _require_saas_storage()
    items = workbook_store.list(user.id)
    # Read the tier straight from public.users so the UI stays in sync when an
    # admin flips the tier manually (until Stripe is live).
    tier = "free"
    try:
        summary = cloud_usage.get_tier_and_usage(user.id)
        tier = summary.get("tier") or "free"
    except Exception:
        pass
    limit = cloud_config.max_workbooks(tier)
    return {
        "workbooks": items,
        "tier": tier,
        "used": len(items),
        "limit": limit,
        "remaining": None if limit == 0 else max(0, limit - len(items)),
    }


@app.post("/workbooks")
async def create_workbook(
    req: WorkbookCreateRequest,
    user: AuthUser = Depends(require_user),
):
    """Create an empty workbook slot. Enforces the per-tier cap; returns 402
    with the current usage so the UI can prompt for upgrade."""
    _require_saas_storage()
    tier = "free"
    try:
        summary = cloud_usage.get_tier_and_usage(user.id)
        tier = summary.get("tier") or "free"
    except Exception:
        pass
    limit = cloud_config.max_workbooks(tier)
    current = workbook_store.count(user.id)
    if limit > 0 and current >= limit:
        raise HTTPException(status_code=402, detail={
            "message": f"Workbook slot cap reached ({current}/{limit}) for your tier.",
            "usage": {"tier": tier, "used": current, "limit": limit},
        })
    created = workbook_store.create_empty(user.id, req.title or "Untitled workbook")
    return {"status": "Success", **created}


@app.patch("/workbooks/{workbook_id}")
async def rename_workbook(
    workbook_id: str,
    req: WorkbookRenameRequest,
    user: AuthUser = Depends(require_user),
):
    _require_saas_storage()
    scope = _scope_for(user, workbook_id)
    workbook_store.rename(scope, req.title)
    return {"status": "Success", "workbook_id": workbook_id, "title": req.title.strip()[:120]}


@app.delete("/workbooks/{workbook_id}")
async def delete_workbook(
    workbook_id: str,
    user: AuthUser = Depends(require_user),
):
    _require_saas_storage()
    scope = _scope_for(user, workbook_id)
    workbook_store.delete(scope)
    return {"status": "Success", "workbook_id": workbook_id}


@app.get("/system/export")
async def export_workbook(k: GridOSKernel = Depends(current_kernel_dep)):
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    safe_name = "".join(c if c.isalnum() or c in "-_ " else "-" for c in kernel.workbook_name).strip() or "workbook"
    safe_name = safe_name.replace(" ", "_")
    filename = f"{safe_name}-{timestamp}.gridos"
    body = json.dumps(kernel.export_state_dict(), indent=2)
    return Response(
        content=body,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/system/export.xlsx")
async def export_workbook_xlsx(k: GridOSKernel = Depends(current_kernel_dep)):
    """Serialize the current workbook to .xlsx. One Excel sheet per GridOS
    sheet (respecting sheet_order). Cells with a `formula` get the formula
    string (Excel recomputes on open for compatible functions); cells without
    get their rendered value. Charts, macros, locked metadata, and the chat
    log are dropped — they have no clean Excel analogue. For a full-fidelity
    export users should use the .gridos format instead."""
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except ImportError:
        raise HTTPException(
            status_code=503,
            detail="openpyxl is not installed. Run `pip install openpyxl`.",
        )

    state = kernel.export_state_dict()
    wb = Workbook()
    # Remove the default sheet; we'll add our own (and openpyxl requires at
    # least one sheet, so we create the first one immediately).
    wb.remove(wb.active)

    sheet_order = state.get("sheet_order") or list((state.get("sheets") or {}).keys())
    if not sheet_order:
        sheet_order = ["Sheet1"]

    sheets_data = state.get("sheets") or {}

    for sheet_name in sheet_order:
        # Excel sheet names can't exceed 31 chars and can't contain []:*?/\.
        safe_sheet = sheet_name[:31]
        for ch in "[]:*?/\\":
            safe_sheet = safe_sheet.replace(ch, "_")
        ws = wb.create_sheet(title=safe_sheet or "Sheet")
        cells = (sheets_data.get(sheet_name) or {}).get("cells") or {}
        for a1, cell in cells.items():
            if not isinstance(cell, dict):
                continue
            formula = cell.get("formula")
            value = cell.get("value")
            datatype = cell.get("datatype")
            if formula:
                ws[a1] = formula  # openpyxl writes leading "=" strings as formulas
            elif datatype == "num" and value not in (None, ""):
                try:
                    ws[a1] = float(value)
                except (TypeError, ValueError):
                    ws[a1] = value
            else:
                ws[a1] = value

    # Ensure we always ship at least one sheet even if the kernel was empty.
    if not wb.worksheets:
        wb.create_sheet(title="Sheet1")

    # openpyxl .save writes to a file-like; stream into an in-memory buffer.
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    safe_name = "".join(c if c.isalnum() or c in "-_ " else "-" for c in kernel.workbook_name).strip() or "workbook"
    safe_name = safe_name.replace(" ", "_")
    filename = f"{safe_name}-{timestamp}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/system/import.xlsx")
async def import_workbook_xlsx(
    file: UploadFile = File(...),
    k: GridOSKernel = Depends(current_kernel_dep),
):
    """Parse a .xlsx upload into GridOS state and apply to the kernel.
    One worksheet → one sheet. Cell values carry over; formula strings are
    preserved so GridOS's evaluator can recompute on next recalc (anything
    GridOS doesn't understand will render as #ERROR in that cell rather
    than silently drop). Styles, merged cells, named ranges, and charts
    are dropped — .gridos round-trips are the lossless path."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        raise HTTPException(
            status_code=503,
            detail="openpyxl is not installed. Run `pip install openpyxl`.",
        )
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Expected an .xlsx file.")
    try:
        raw = await file.read()
        from io import BytesIO
        wb = load_workbook(BytesIO(raw), data_only=False)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse Excel file: {e}")

    sheets: dict[str, dict] = {}
    sheet_order: list[str] = []
    for ws in wb.worksheets:
        name = ws.title or "Sheet"
        sheet_order.append(name)
        cells: dict[str, dict] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                raw_val = cell.value
                formula = None
                value_str = ""
                datatype = "str"
                if isinstance(raw_val, str) and raw_val.startswith("="):
                    formula = raw_val
                    # Value will be filled by GridOS's evaluator on next recalc.
                elif isinstance(raw_val, (int, float)):
                    datatype = "num"
                    value_str = str(raw_val)
                elif isinstance(raw_val, bool):
                    value_str = "TRUE" if raw_val else "FALSE"
                else:
                    value_str = str(raw_val)
                cells[cell.coordinate] = {
                    "value": value_str,
                    "formula": formula,
                    "locked": False,
                    "datatype": datatype,
                    "agent_owner": "User",
                }
        sheets[name] = {"cells": cells, "charts": []}

    if not sheet_order:
        sheet_order = ["Sheet1"]
        sheets["Sheet1"] = {"cells": {}, "charts": []}

    # Derive a friendly workbook name from the uploaded filename minus .xlsx.
    base = file.filename.rsplit(".", 1)[0]
    workbook_name = (base or "Imported workbook").strip()[:120] or "Imported workbook"

    state = {
        "workbook_name": workbook_name,
        "active_sheet": sheet_order[0],
        "sheet_order": sheet_order,
        "sheets": sheets,
        "chat_log": [],
    }
    try:
        kernel.apply_state_dict(state)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not apply imported state: {e}")
    return {"status": "Success", "workbook_name": workbook_name, "sheets": len(sheet_order)}


@app.post("/system/import")
async def import_workbook(
    payload: dict = Body(...),
    k: GridOSKernel = Depends(current_kernel_dep),
):
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Invalid workbook payload.")
    try:
        kernel.apply_state_dict(payload)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not import workbook: {e}")
    return {"status": "Success"}


@app.post("/formula/evaluate")
async def evaluate_formula(req: FormulaRequest):
    evaluator = FormulaEvaluator()
    return {"result": evaluator.evaluate(req.function_name, req.arguments)}


@app.post("/system/clear")
async def clear_grid(
    sheet: Optional[str] = None,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    kernel.clear_unlocked(sheet)
    return {"status": "Success", "sheet": sheet or kernel.active_sheet}


@app.post("/system/unlock-all")
async def unlock_all(k: GridOSKernel = Depends(current_kernel_dep)):
    """Forcibly clear the locked flag on every cell across every sheet, and drop
    any empty-locked placeholder cells that have no value/formula."""
    dropped = 0
    unlocked = 0
    for sheet_name, state in kernel.sheets.items():
        cells = state["cells"]
        for coords in list(cells.keys()):
            cell = cells[coords]
            if not cell.locked:
                continue
            if cell.value in (None, "") and not cell.formula:
                del cells[coords]
                dropped += 1
            else:
                cell.locked = False
                unlocked += 1
    return {"status": "Success", "unlocked": unlocked, "dropped": dropped}


# ---------- Charts ----------


class ChartCreateRequest(BaseModel):
    anchor_cell: str = "F2"
    data_range: str
    chart_type: str = "bar"
    title: str = ""
    width: int = 400
    height: int = 280
    orientation: str = "columns"
    sheet: Optional[str] = None


class ChartUpdateRequest(BaseModel):
    anchor_cell: Optional[str] = None
    data_range: Optional[str] = None
    chart_type: Optional[str] = None
    title: Optional[str] = None
    width: Optional[int] = None
    height: Optional[int] = None
    orientation: Optional[str] = None
    sheet: Optional[str] = None


@app.get("/system/charts")
async def list_charts(
    sheet: Optional[str] = None,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    return {"charts": kernel.list_charts(sheet)}


@app.post("/system/charts")
async def create_chart(
    req: ChartCreateRequest,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    spec = req.model_dump(exclude={"sheet"})
    try:
        chart = kernel.add_chart(spec, sheet_name=req.sheet)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not create chart: {e}")
    return {"status": "Success", "chart": chart}


@app.patch("/system/charts/{chart_id}")
async def update_chart(
    chart_id: str,
    req: ChartUpdateRequest,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    updates = req.model_dump(exclude={"sheet"}, exclude_none=True)
    try:
        chart = kernel.update_chart(chart_id, updates, sheet_name=req.sheet)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not update chart: {e}")
    return {"status": "Success", "chart": chart}


@app.delete("/system/charts/{chart_id}")
async def delete_chart(
    chart_id: str,
    sheet: Optional[str] = None,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    if not kernel.delete_chart(chart_id, sheet_name=sheet):
        raise HTTPException(status_code=404, detail=f"Chart '{chart_id}' not found.")
    return {"status": "Success"}


# ---------- Library: templates ----------


_TEMPLATE_ID_RE = re.compile(r"^[A-Za-z0-9_\-]+$")
_SAFE_SLUG_RE = re.compile(r"[^A-Za-z0-9_\-]+")


def _slugify_template_name(name: str) -> str:
    slug = _SAFE_SLUG_RE.sub("-", name.strip().lower()).strip("-")
    return slug or "template"


def _template_path(template_id: str) -> Path:
    if not _TEMPLATE_ID_RE.match(template_id):
        raise HTTPException(status_code=400, detail="Invalid template id.")
    return TEMPLATES_DIR / f"{template_id}.json"


class TemplateSaveRequest(BaseModel):
    name: str
    description: Optional[str] = ""


class MacroSaveRequest(BaseModel):
    name: str
    description: Optional[str] = ""
    params: List[str] = []
    body: str


class HeroToolToggleRequest(BaseModel):
    tool_id: str
    enabled: bool


@app.post("/templates/save")
async def save_template(
    req: TemplateSaveRequest,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    if not req.name or not req.name.strip():
        raise HTTPException(status_code=400, detail="Template name is required.")

    base_slug = _slugify_template_name(req.name)
    candidate = base_slug
    counter = 2
    while (TEMPLATES_DIR / f"{candidate}.json").exists():
        candidate = f"{base_slug}-{counter}"
        counter += 1

    created_at = datetime.now(timezone.utc).isoformat()
    snapshot = {
        "id": candidate,
        "name": req.name.strip(),
        "description": (req.description or "").strip(),
        "author": "You",
        "created_at": created_at,
        "state": kernel.export_state_dict(),
    }
    (TEMPLATES_DIR / f"{candidate}.json").write_text(
        json.dumps(snapshot, indent=2), encoding="utf-8"
    )
    return {"status": "Success", "template": _template_summary(snapshot)}


def _template_summary(payload: dict) -> dict:
    state = payload.get("state") or {}
    sheets = state.get("sheets") or {}
    cell_count = 0
    for sheet in sheets.values():
        cell_count += len((sheet or {}).get("cells") or {})
    return {
        "id": payload.get("id"),
        "name": payload.get("name"),
        "description": payload.get("description", ""),
        "author": payload.get("author") or "You",
        "created_at": payload.get("created_at"),
        "sheet_count": len(sheets),
        "cell_count": cell_count,
    }


@app.get("/templates/list")
async def list_templates():
    templates: list[dict] = []
    for path in sorted(TEMPLATES_DIR.glob("*.json")):
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue
        templates.append(_template_summary(payload))
    templates.sort(key=lambda t: t.get("created_at") or "", reverse=True)
    return {"templates": templates}


@app.get("/templates/load/{template_id}")
async def load_template(template_id: str):
    path = _template_path(template_id)
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"Template '{template_id}' not found.")
    return json.loads(path.read_text(encoding="utf-8"))


@app.post("/templates/apply/{template_id}")
async def apply_template(
    template_id: str,
    k: GridOSKernel = Depends(current_kernel_dep),
):
    path = _template_path(template_id)
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"Template '{template_id}' not found.")
    payload = json.loads(path.read_text(encoding="utf-8"))
    try:
        result = kernel.apply_template_respecting_locks(payload.get("state") or {})
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not apply template: {e}")
    return {"status": "Success", **result}


@app.delete("/templates/{template_id}")
async def delete_template(template_id: str):
    path = _template_path(template_id)
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"Template '{template_id}' not found.")
    path.unlink()
    return {"status": "Success"}


# ---------- Settings: API keys + model catalog ----------


def _mask_key(key: str) -> str:
    if not key:
        return ""
    if len(key) <= 8:
        return "•" * len(key)
    return f"{key[:4]}…{key[-4:]}"


@app.get("/settings/providers")
async def list_providers():
    """Describe every known provider and whether a key is configured for it."""
    providers = []
    for pid, cls in PROVIDER_CLASSES.items():
        key = API_KEYS.get(pid, "")
        providers.append({
            "id": pid,
            "display_name": PROVIDER_DISPLAY_NAMES.get(pid, pid),
            "configured": bool(key and pid in PROVIDERS),
            "masked_key": _mask_key(key) if key else "",
        })
    return {"providers": providers}


@app.post("/settings/keys/save")
async def save_api_key(req: ApiKeySaveRequest):
    provider_id = (req.provider or "").strip().lower()
    cls = PROVIDER_CLASSES.get(provider_id)
    if cls is None:
        raise HTTPException(status_code=400, detail=f"Unknown provider: {req.provider}")
    api_key = (req.api_key or "").strip()
    if not api_key:
        raise HTTPException(status_code=400, detail="API key is empty.")

    # Try to instantiate the provider before persisting — catches missing SDKs
    # (e.g. `anthropic` not installed) and gives the user an actionable message.
    try:
        cls(api_key=api_key)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not initialize {provider_id}: {e}")

    API_KEYS[provider_id] = api_key
    _persist_api_keys()
    _rebuild_providers()
    return {
        "status": "Success",
        "provider": provider_id,
        "configured": provider_id in PROVIDERS,
    }


@app.delete("/settings/keys/{provider_id}")
async def delete_api_key(provider_id: str):
    provider_id = (provider_id or "").strip().lower()
    if provider_id not in PROVIDER_CLASSES:
        raise HTTPException(status_code=400, detail=f"Unknown provider: {provider_id}")
    if provider_id in API_KEYS:
        del API_KEYS[provider_id]
        _persist_api_keys()
    _rebuild_providers()
    return {"status": "Success", "provider": provider_id}


@app.get("/models/available")
async def list_available_models():
    """Every model whose provider currently has a working key. The UI uses this to
    populate the per-chat model picker."""
    configured = _configured_provider_ids()
    default_id = default_model_id(configured)
    models = [
        {**entry, "available": entry["provider"] in configured}
        for entry in MODEL_CATALOG
    ]
    return {
        "models": models,
        "default_model_id": default_id,
        "configured_providers": sorted(configured),
    }


# ---------- Library: tools ----------


@app.get("/tools/list")
async def list_tools():
    return {
        "primitives": [
            {"name": name, "builtin": True}
            for name in _builtin_primitive_names()
            if name.upper() not in _macro_names()
        ],
        "macros": [dict(m) for m in USER_MACROS],
        "hero_tools": [
            {
                "id": t["id"],
                "display_name": t["display_name"],
                "description": t["description"],
                "enabled": bool(HERO_TOOLS_STATE.get(t["id"], False)),
            }
            for t in HERO_TOOLS_CATALOG
        ],
    }


@app.post("/tools/save_macro")
async def save_macro(req: MacroSaveRequest):
    clean_name = (req.name or "").strip()
    if not clean_name:
        raise HTTPException(status_code=400, detail="Macro name is required.")
    clean_body = (req.body or "").strip()
    if not clean_body:
        raise HTTPException(status_code=400, detail="Macro body is required.")

    spec = {
        "name": clean_name,
        "description": (req.description or "").strip(),
        "params": list(req.params or []),
        "body": clean_body,
    }

    try:
        _register_macro(spec)
    except MacroError as e:
        raise HTTPException(status_code=400, detail=str(e))

    normalized = {
        "name": clean_name.upper(),
        "description": spec["description"],
        "params": [p.upper() for p in spec["params"]],
        "body": clean_body,
    }
    replaced = False
    for idx, existing in enumerate(USER_MACROS):
        if existing["name"] == normalized["name"]:
            USER_MACROS[idx] = normalized
            replaced = True
            break
    if not replaced:
        USER_MACROS.append(normalized)

    _persist_user_macros()
    return {"status": "Success", "macro": normalized, "replaced": replaced}


@app.delete("/tools/macros/{macro_name}")
async def delete_macro(macro_name: str):
    upper = macro_name.upper()
    removed = False
    for idx, existing in enumerate(list(USER_MACROS)):
        if existing["name"] == upper:
            USER_MACROS.pop(idx)
            removed = True
            break
    if not removed:
        raise HTTPException(status_code=404, detail=f"Macro '{macro_name}' not found.")
    _unregister_macro(upper)
    _persist_user_macros()
    return {"status": "Success"}


@app.post("/tools/hero/toggle")
async def toggle_hero_tool(req: HeroToolToggleRequest):
    if req.tool_id not in HERO_TOOLS_STATE:
        raise HTTPException(status_code=404, detail=f"Unknown hero tool '{req.tool_id}'.")
    HERO_TOOLS_STATE[req.tool_id] = bool(req.enabled)
    _persist_hero_tools()
    return {"status": "Success", "tool_id": req.tool_id, "enabled": HERO_TOOLS_STATE[req.tool_id]}
